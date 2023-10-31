import requests
from openpyxl import Workbook, load_workbook
import os

if not os.path.exists("kommunalskatt.xlsx"):
    excel = Workbook()
    sheet = excel.active
    sheet.title = "Kommunal Skatt Data"
    sheet["B1"] = '2014'
    sheet["C1"] = '2015'
    sheet['D1'] = '2016'
    sheet['E1'] = '2017'
    sheet['F1'] = '2018'
    sheet['G1'] = '2019'
    sheet['H1'] = '2020'
    sheet['I1'] = '2021'
    sheet['J1'] = '2022'
    sheet['K1'] = '2023'
    excel.save('kommunalskatt.xlsx')



url = "https://skatteverket.entryscape.net/rowstore/dataset/c67b320b-ffee-4876-b073-dd9236cd2a99"
år_att_kolla = range(2014, 2024)
stockholms_län = ["BOTKYRKA"]#,"DANDERYD","EKERÖ","HANINGE","HUDDINGE","JÄRFÄLLA","LIDINGÖ","NACKA","NORRTÄLJE","NYKVARN","NYNÄSHAMN","SALEM","SIGTUNA","SOLLENTUNA","SOLNA","STOCKHOLM","SUNDBYBERG","SÖDERTÄLJE","TYRESÖ","TÄBY","UPPLANDS VÄSBY","UPPLANDS-BRO","VALLENTUNA","VAXHOLM","VÄRMDÖ","ÖSTERÅKER"]

for kommun in stockholms_län:

    # Lägg till kommunens namn i första platsen i en lista
    kommun_data = [kommun]

    for år in år_att_kolla:
        år_url = f"{url}?år={år}"
        response = requests.get(år_url)

        if response.status_code == 200:
            data = response.json()
            snitt_skatt = []
            
            for entry in data['results']:

                if entry['år'] == str(år) and entry['kommun'] == kommun:
                    snitt_skatt.append(float(entry['kommunal-skatt']))
                    
        # Ett snitt räknas ut för de tillfällen där kommunen har flera församlingar
        skatt = sum(snitt_skatt) / len(snitt_skatt)
        kommun_data.append(skatt)

        # Lägg till listan på nästa rad i Excel-filen
        excel = load_workbook('kommunalskatt.xlsx')
        excel.append(kommun_data)
        excel.save('kommunalskatt.xlsx')



